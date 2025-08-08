import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import unicodedata 
# Set page config to wide mode
st.set_page_config(
    page_title="Excel Qiymətlər",
    layout="wide"
)

# Your desired translations
translated_drag_and_drop = "Qiymətlər olan Exceli bura yükləyin"
translated_limit = " "
translated_browse_files = "Faylı seç"

# Custom CSS to make it even wider if needed and fix multiselect
st.markdown(
    """
    <style>
    .main .block-container {
        max-width: 95%;
        padding-left: 1rem;
        padding-right: 1rem;
    }
    
    [data-testid="stFileUploaderDropzone"] div div::before {
        content: '""" + translated_drag_and_drop + """';
        color: #6C757D;
    }
    [data-testid="stFileUploaderDropzone"] div div span {
        display: none;
    }
    [data-testid="stFileUploaderDropzone"] div div::after {
        content: '""" + translated_limit + """';
        color: #6C757D;
    }
    [class="st-emotion-cache-r92n3i e7nj0r42"] {
        font-size: 0;
        position: relative;
    }
    [class="st-emotion-cache-r92n3i e7nj0r42"]::after {
        content: '""" + translated_browse_files + """';
        font-size: 1rem;
        visibility: visible;
    }
   
    .stMultiSelect,
    .stMultiSelect > div > div,
    .stMultiSelect [data-baseweb="select"],
    .stMultiSelect [data-baseweb="popover"] {
        width: 100% !important;
        min-width: 100% !important;
    }

    .stMultiSelect [data-baseweb="select"],
    .stMultiSelect [data-baseweb="select"] > div,
    .stMultiSelect [data-baseweb="select"] [data-baseweb="input"] {
        min-height: 38px !important;
    }

    .stMultiSelect [data-baseweb="select"] > div {
        flex-wrap: wrap !important;
    }

    .stMultiSelect [data-baseweb="tag"],
    .stMultiSelect [data-baseweb="tag"] span {
        max-width: none !important;
        white-space: nowrap !important;
        overflow: visible !important;
        text-overflow: clip !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.title("Excel Qiymətlər")

uploaded_file = st.file_uploader("Exceli yüklə", type=["xlsx"])

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    ws = wb.active
    
    cols_to_copy = [4, 7, 8, 13, 14, 15]
    
    title_cell = ws.cell(row=1, column=4).value
    # st.write("Title from D1:", title_cell)
    
    raw_headers = [ws.cell(row=2, column=col).value for col in cols_to_copy]
    
    selected_headers = []
    for i, header in enumerate(raw_headers):
        if header is None or header == "":
            col_letter = openpyxl.utils.get_column_letter(cols_to_copy[i])
            unique_header = f"Column_{col_letter}"
        else:
            unique_header = str(header).strip()
        
        original_header = unique_header
        counter = 1
        while unique_header in selected_headers:
            unique_header = f"{original_header}_{counter}"
            counter += 1
        
        selected_headers.append(unique_header)
    
    assignments_col_index = None
    assignments_excel_col = None
    for i, header in enumerate(raw_headers):
        if header and "assignment" in str(header).lower():
            assignments_col_index = i
            assignments_excel_col = cols_to_copy[i]
            break
    
    data = []
    filtered_rows_count = 0
    total_rows_count = 0
    
    for row in range(3, ws.max_row + 1):
        total_rows_count += 1
        row_values = [ws.cell(row=row, column=col).value for col in cols_to_copy]
        
        if assignments_col_index is not None:
            assignments_value = row_values[assignments_col_index]
            if assignments_value and isinstance(assignments_value, str):
                if "riyaziyyat" in assignments_value.lower():
                    data.append(row_values)
                    filtered_rows_count += 1
        else:
            data.append(row_values)
    
    df = pd.DataFrame(data, columns=selected_headers)
    
    assignments_col = None
    for col_name in selected_headers:
        if col_name and "assignment" in col_name.lower():
            assignments_col = col_name
            break
    
    if assignments_col is None:
        st.error("'Assignments' sütunu tapılmadı")
        st.write("sütunlar:", selected_headers)
    else:
        assignments_series = df[assignments_col].dropna()
        assignments_series = assignments_series[assignments_series != ""]
        assignments_options = sorted(assignments_series.astype(str).unique())
        
        if len(assignments_options) == 0:
            st.warning("İmtahan tapılmadı sütunda")
        else:
            selected_assignments = st.multiselect(
                "İmtahanları seç",
                assignments_options,
                placeholder="İmtahanları seçin",
                help="Bir və ya daha çox imtahan seç"
            )
            
            if selected_assignments:
                mask = df[assignments_col].astype(str).isin(selected_assignments)
                filtered_df = df[mask]
                st.write(f"Seçilmiş ({len(filtered_df)} nəticələr):")
                
                try:
                    display_df = filtered_df.copy()
                    for col in display_df.columns:
                        if display_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                            numeric_vals = display_df[col].dropna()
                            if len(numeric_vals) > 0 and numeric_vals.min() >= 0 and numeric_vals.max() <= 1:
                                display_df[col] = display_df[col].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else x)
                    
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        height=1200
                    )
                except Exception as e:
                    st.error(f"Error displaying filtered data: {e}")
                
                email_col = None
                for col_name in selected_headers:
                    if col_name and "email" in col_name.lower():
                        email_col = col_name
                        break
                
                if email_col is None:
                    st.warning("Email ünvan sütunu yoxdur")
                    allow_download = False
                else:
                    duplicates_mask = filtered_df.duplicated(subset=[email_col], keep=False)
                    duplicated_df = filtered_df[duplicates_mask]
                    
                    if len(duplicated_df) > 0:
                        st.warning(f"⚠️ {len(duplicated_df)} dənə eyni imtahan nəticəsi olan şagird tapıldı")
                        
                        duplicate_groups = duplicated_df.groupby(email_col)
                        
                        st.subheader("Eyni şagirdlərin yalnız bir nəticəsin seçin")
                        
                        if 'selected_duplicates' not in st.session_state:
                            st.session_state.selected_duplicates = {}
                        
                        final_df = filtered_df[~duplicates_mask].copy()
                        all_selected = True
                        
                        for email, group in duplicate_groups:
                            st.write(f"**Email: {email}** ({len(group)} təkrar)")
                            
                            cols = st.columns(len(group))
                            
                            for i, (idx, row) in enumerate(group.iterrows()):
                                with cols[i]:
                                    is_selected = st.session_state.selected_duplicates.get(email) == idx
                                    
                                    summary_parts = []
                                    
                                    for col in group.columns[:3]:
                                        val = row[col]
                                        if val is not None and isinstance(val, (int, float)) and 0 <= val <= 1:
                                            formatted_val = f"{val * 100:.1f}%"
                                        else:
                                            formatted_val = str(val)[:20] if val is not None else "None"
                                        summary_parts.append(f"**{col}:** {formatted_val}")
                                    
                                    points_col = None
                                    for col_name in group.columns:
                                        if col_name and "point" in col_name.lower():
                                            points_col = col_name
                                            break
                                    
                                    if points_col and points_col not in group.columns[:3]:
                                        points_val = row[points_col]
                                        if points_val is not None and isinstance(points_val, (int, float)) and 0 <= points_val <= 1:
                                            formatted_points = f"{points_val * 100:.1f}%"
                                        else:
                                            formatted_points = str(points_val) if points_val is not None else "None"
                                        summary_parts.append(f"**{points_col}:** {formatted_points}")
                                    
                                    summary = "\n\n".join(summary_parts)
                                    
                                    if is_selected:
                                        st.success(f"✅ **SEÇİLDİ**\n\n**Sıra {idx}**\n\n{summary}")
                                    else:
                                        st.error(f"❌ **Sıra {idx}**\n\n{summary}")
                                    
                                    if st.button(f"Seç Sıra {idx}", key=f"select_{email}_{idx}"):
                                        st.session_state.selected_duplicates[email] = idx
                                        st.rerun()
                            
                            if email not in st.session_state.selected_duplicates:
                                all_selected = False
                            else:
                                selected_idx = st.session_state.selected_duplicates[email]
                                final_df = pd.concat([final_df, group.loc[[selected_idx]]])
                        
                        if all_selected:
                            st.success("✅ Təkrarlanan şagird adı yoxdur. Yükləyə bilərsiz!")
                            allow_download = True
                            final_filtered_df = final_df.reset_index(drop=True)
                            final_filtered_df.index = final_filtered_df.index + 1
                            
                            final_display_df = final_filtered_df.copy()
                            for col in final_display_df.columns:
                                if final_display_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                                    numeric_vals = final_display_df[col].dropna()
                                    if len(numeric_vals) > 0 and numeric_vals.min() >= 0 and numeric_vals.max() <= 1:
                                        final_display_df[col] = final_display_df[col].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else x)
                            
                            st.write(f"Final data ({len(final_filtered_df)} rows after removing duplicates):")
                            st.dataframe(final_display_df, use_container_width=True, height=400)
                        else:
                            st.error("❌ Yükləməzdən əvvəl təkrarları düzəldin")
                            allow_download = False
                            final_filtered_df = filtered_df
                    else:
                        st.success("✅ Təkralanan şagird tapılmadı")
                        allow_download = True
                        final_filtered_df = filtered_df
            else:
                filtered_df = df
                st.write(f"Bütün ({len(filtered_df)} sıralar):")
                
                try:
                    display_df = filtered_df.copy()
                    for col in display_df.columns:
                        if display_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                            numeric_vals = display_df[col].dropna()
                            if len(numeric_vals) > 0 and numeric_vals.min() >= 0 and numeric_vals.max() <= 1:
                                display_df[col] = display_df[col].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else x)
                    
                    st.dataframe(
                        display_df,
                        use_container_width=True,
                        height=1200
                    )
                except Exception as e:
                    st.error(f"Error displaying filtered data: {e}")
                
                allow_download = True
                final_filtered_df = filtered_df
            
            # Excel download function
            def to_excel(df, title):
                output = BytesIO()
                try:
                    excel_df = df.copy()
                    percentage_columns = []
                    
                    for col in excel_df.columns:
                        if excel_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                            numeric_vals = excel_df[col].dropna()
                            if len(numeric_vals) > 0 and numeric_vals.min() >= 0 and numeric_vals.max() <= 1:
                                percentage_columns.append(col)
                    
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        excel_df.to_excel(writer, index=False, sheet_name='FilteredData', startrow=1, startcol=0)
                        
                        workbook = writer.book
                        worksheet = writer.sheets['FilteredData']

                        content_font = Font(name='Segoe UI')
                        header_fill = PatternFill(start_color='5B5FC7', end_color='5B5FC7', fill_type='solid')

                        light_gray_fill = PatternFill(start_color='E7E7F7', end_color='E7E7F7', fill_type='solid')
                        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                        for row in worksheet.iter_rows():
                            for cell in row:
                                cell.font = content_font
                        
                        for row in worksheet.iter_rows(min_row=1, max_row=2):
                            for cell in row:
                                cell.fill = header_fill

                        header_row = worksheet[2]
                        for cell in header_row:
                            cell.font = Font(name='Segoe UI', bold=True, color='FFFFFF')

                        if title:
                            title_cell_obj = worksheet.cell(row=1, column=1, value=title)
                            title_cell_obj.font = Font(name='Segoe UI', size=18, bold=True, color='FFFFFF')
                         
                        for row_index, row in enumerate(worksheet.iter_rows(min_row=3, max_row=len(excel_df) + 2)):
                            fill = light_gray_fill if row_index % 2 == 0 else white_fill
                            for cell in row:
                                cell.fill = fill

                        manual_widths = {'A': 30, 'C': 35}

                        for i, column_name in enumerate(excel_df.columns):
                            column_letter = get_column_letter(i + 1)
                            
                            if column_letter in manual_widths:
                                worksheet.column_dimensions[column_letter].width = manual_widths[column_letter]
                            else:
                                max_length = 0
                                for cell in worksheet[column_letter]:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = (max_length + 2)
                                worksheet.column_dimensions[column_letter].width = adjusted_width

                        filter_range = f'A2:{get_column_letter(len(excel_df.columns))}2'
                        worksheet.auto_filter.ref = filter_range

                        if percentage_columns:
                            for col_name in percentage_columns:
                                try:
                                    col_idx = list(excel_df.columns).index(col_name) + 1
                                    for row in range(3, len(excel_df) + 3):
                                        cell = worksheet.cell(row=row, column=col_idx)
                                        if cell.value is not None:
                                            cell.number_format = '0.0%'
                                except Exception as col_error:
                                    print(f"Error formatting column {col_name}: {col_error}")
                    
                    output.seek(0)
                    return output
                except Exception as e:
                    st.error(f"Error creating Excel file: {e}")
                    return None
            # PDF download function
            def to_pdf(df, title):
                output = BytesIO()
                try:
                    doc = SimpleDocTemplate(output, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch,
                                          leftMargin=0.5*inch, rightMargin=0.5*inch)

                    styles = getSampleStyleSheet()

                    try:
                        pdfmetrics.registerFont(TTFont('NotoSans-Regular', 'fonts/NotoSans-Regular.ttf'))
                        pdfmetrics.registerFont(TTFont('NotoSans-Bold', 'fonts/NotoSans-Bold.ttf'))
                        font_name = 'NotoSans-Regular'
                        font_name_bold = 'NotoSans-Bold'
                    except Exception as e:
                        st.error(f"Noto Sans fonts could not be loaded: {e}")
                        st.error("Make sure fonts/NotoSans-Regular.ttf and fonts/NotoSans-Bold.ttf exist")
                        return None

                    title_style = ParagraphStyle(
                        'CustomTitle',
                        fontName=font_name_bold,
                        parent=styles['Heading1'],
                        fontSize=16,
                        spaceAfter=20,
                        alignment=1,
                        textColor=colors.HexColor('#5B5FC7')
                    )

                    story = []
                    if title:
                        title_text = str(title) if title else ""
                        title_para = Paragraph(title_text, title_style)
                        story.append(title_para)
                        story.append(Spacer(1, 12))

                    pdf_df = df.copy()

                    # --- MODIFICATION START: Direct character removal ---
                    # This is the most reliable way to fix the double-dot 'i' problem.
                    # It removes the combining dot character U+0307 from all string columns.
                    problematic_char = '\u0307' # Unicode for COMBINING DOT ABOVE
                    for col in pdf_df.columns:
                        if pd.api.types.is_string_dtype(pdf_df[col]):
                            pdf_df[col] = pdf_df[col].astype(str).str.replace(problematic_char, '', regex=False)
                    # --- MODIFICATION END ---
                    
                    first_col_index = 0
                    long_content_col_indices = []
                    short_content_col_indices = []
                    
                    for i, col in enumerate(pdf_df.columns):
                        col_name_lower = str(col).lower()
                        if i == first_col_index:
                            pass
                        elif "assignment" in col_name_lower or "email" in col_name_lower:
                            long_content_col_indices.append(i)
                        else:
                            short_content_col_indices.append(i)

                    for col in pdf_df.columns:
                        if pdf_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                            numeric_vals = pdf_df[col].dropna()
                            if len(numeric_vals) > 0 and numeric_vals.min() >= 0 and numeric_vals.max() <= 1:
                                pdf_df[col] = pdf_df[col].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else "")
                    
                    first_col_style = ParagraphStyle(
                        'FirstColStyle',
                        fontName=font_name,
                        fontSize=9,
                        alignment=0,
                        leading=11
                    )
                    long_col_style = ParagraphStyle(
                        'LongColStyle',
                        fontName=font_name,
                        fontSize=8,
                        alignment=0,
                        leading=10
                    )
                    
                    data = []
                    headers = [str(col) if col is not None else "" for col in pdf_df.columns]
                    data.append(headers)

                    for _, row in pdf_df.iterrows():
                        row_data = []
                        for i, val in enumerate(row):
                            cell_text = str(val) if val is not None else ""
                            
                            if i == first_col_index:
                                wrapped_text = Paragraph(cell_text, first_col_style)
                                row_data.append(wrapped_text)
                            elif i in long_content_col_indices and len(cell_text) > 20:
                                wrapped_text = Paragraph(cell_text, long_col_style)
                                row_data.append(wrapped_text)
                            else:
                                row_data.append(cell_text)
                        data.append(row_data)

                    page_width = A4[0] - 2 * 0.5 * inch
                    num_cols = len(pdf_df.columns)
                    
                    col_widths = [0] * num_cols
                    
                    first_col_width = 1.5 * inch
                    col_widths[first_col_index] = first_col_width
                    
                    long_col_width = 1.2 * inch
                    for i in long_content_col_indices:
                        col_widths[i] = long_col_width
                        
                    total_assigned_width = first_col_width + (len(long_content_col_indices) * long_col_width)
                    remaining_width = page_width - total_assigned_width
                    
                    num_short_cols = len(short_content_col_indices)
                    if num_short_cols > 0:
                        short_col_width = remaining_width / num_short_cols
                        for i in short_content_col_indices:
                            col_widths[i] = short_col_width

                    table = Table(data, colWidths=col_widths)

                    table_style = [
                        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                        ('FONTNAME', (0, 1), (-1, -1), font_name),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B5FC7')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('ALIGN', (first_col_index, 1), (first_col_index, -1), 'LEFT'),
                        ('VALIGN', (first_col_index, 1), (first_col_index, -1), 'TOP'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E7E7F7')]),
                    ]

                    for col_index in long_content_col_indices:
                        table_style.append(('ALIGN', (col_index, 1), (col_index, -1), 'LEFT'))
                        table_style.append(('VALIGN', (col_index, 1), (col_index, -1), 'TOP'))
                    
                    for col_index in short_content_col_indices:
                        table_style.append(('ALIGN', (col_index, 1), (col_index, -1), 'CENTER'))
                        table_style.append(('VALIGN', (col_index, 1), (col_index, -1), 'MIDDLE'))

                    table.setStyle(TableStyle(table_style))

                    story.append(table)

                    footer_style = ParagraphStyle(
                        'Footer',
                        fontName=font_name,
                        parent=styles['Normal'],
                        fontSize=8,
                        spaceAfter=12,
                        alignment=1,
                        textColor=colors.grey
                    )

                    story.append(Spacer(1, 20))
                    footer_text = f"Nəticələr sayı: {len(pdf_df)} | Yaradılma tarixi: {pd.Timestamp.now().strftime('%d.%m.%Y %H:%M')}"
                    footer_para = Paragraph(footer_text, footer_style)
                    story.append(footer_para)

                    doc.build(story)
                    output.seek(0)
                    return output

                except Exception as e:
                    st.error(f"PDF yaradılarkən xəta: {e}")
                    return None
            # Add this function after your existing to_pdf function

            def to_pdf_landscape(df, title):
                output = BytesIO()
                try:
                    # Use landscape orientation - swap width and height of A4
                    doc = SimpleDocTemplate(output, pagesize=(A4[1], A4[0]), topMargin=0.5*inch, bottomMargin=0.5*inch,
                                          leftMargin=0.5*inch, rightMargin=0.5*inch)

                    styles = getSampleStyleSheet()

                    try:
                        pdfmetrics.registerFont(TTFont('Segoe UI', 'fonts/randomf.ttf'))
                        pdfmetrics.registerFont(TTFont('Segoe UI-Bold', 'fonts/randomf.ttf'))
                        font_name = 'Segoe UI'
                        font_name_bold = 'Segoe UI-Bold'
                    except Exception as e:
                        st.error(f"Noto Sans fonts could not be loaded: {e}")
                        st.error("Make sure fonts/NotoSans-Regular.ttf and fonts/NotoSans-Bold.ttf exist")
                        return None

                    title_style = ParagraphStyle(
                        'CustomTitle',
                        fontName=font_name_bold,
                        parent=styles['Heading1'],
                        fontSize=16,
                        spaceAfter=20,
                        alignment=1,
                        textColor=colors.HexColor('#5B5FC7')
                    )

                    story = []
                    if title:
                        title_text = str(title) if title else ""
                        title_para = Paragraph(title_text, title_style)
                        story.append(title_para)
                        story.append(Spacer(1, 12))

                    pdf_df = df.copy()

                    # Remove problematic characters
                    problematic_char = '\u0307' # Unicode for COMBINING DOT ABOVE
                    for col in pdf_df.columns:
                        if pd.api.types.is_string_dtype(pdf_df[col]):
                            pdf_df[col] = pdf_df[col].astype(str).str.replace(problematic_char, '', regex=False)
                    
                    # Format percentage columns
                    for col in pdf_df.columns:
                        if pdf_df[col].dtype in ['float64', 'float32', 'int64', 'int32']:
                            numeric_vals = pdf_df[col].dropna()
                            if len(numeric_vals) > 0 and numeric_vals.min() >= 0 and numeric_vals.max() <= 1:
                                pdf_df[col] = pdf_df[col].apply(lambda x: f"{x*100:.1f}%" if pd.notna(x) else "")
                    
                    # Simple paragraph style for all content
                    content_style = ParagraphStyle(
                        'ContentStyle',
                        fontName=font_name,
                        fontSize=10,
                        alignment=0,
                        leading=12
                    )
                    
                    data = []
                    headers = [str(col) if col is not None else "" for col in pdf_df.columns]
                    data.append(headers)

                    # Add data rows - wrap longer content in paragraphs
                    for _, row in pdf_df.iterrows():
                        row_data = []
                        for val in row:
                            cell_text = str(val) if val is not None else ""
                            # Wrap longer text content in paragraphs for better formatting
                            if len(cell_text) > 30:
                                wrapped_text = Paragraph(cell_text, content_style)
                                row_data.append(wrapped_text)
                            else:
                                row_data.append(cell_text)
                        data.append(row_data)

                    # Calculate page width for landscape (A4 height becomes width)
                    page_width = A4[1] - 2 * 0.5 * inch
                    num_cols = len(pdf_df.columns)
                    
                    # Smart column width distribution for landscape
                    col_widths = [0] * num_cols
                    
                    # Identify column types based on names and content
                    first_col_index = 0         # First column (likely full names) - needs extra width
                    short_numeric_cols = []     # Points, Max Points, Percent - these need less width
                    long_content_cols = []      # Columns with longer text content
                    regular_cols = []           # Everything else
                    
                    for i, col_name in enumerate(pdf_df.columns):
                        col_name_lower = str(col_name).lower()
                        if i == first_col_index:
                            # First column handled separately
                            continue
                        elif any(keyword in col_name_lower for keyword in ['point', 'percent', 'max']):
                            short_numeric_cols.append(i)
                        elif i == 2 or any(keyword in col_name_lower for keyword in ['assignment', 'email']):
                            # Third column (index 2) or columns with typically longer content
                            long_content_cols.append(i)
                        else:
                            regular_cols.append(i)
                    
                    # Allocate widths: first column gets extra width, short numeric cols get less, long content cols get more
                    first_col_width = 1.8 * inch      # Extra width for full names
                    short_col_width = 0.8 * inch      # Narrow columns for numeric data
                    regular_col_width = 1.2 * inch    # Regular width
                    
                    # Calculate remaining width for long content columns
                    used_width = (first_col_width + 
                                 len(short_numeric_cols) * short_col_width + 
                                 len(regular_cols) * regular_col_width)
                    remaining_width = page_width - used_width
                    
                    if len(long_content_cols) > 0:
                        long_col_width = remaining_width / len(long_content_cols)
                    else:
                        long_col_width = regular_col_width
                    
                    # Assign widths
                    for i in range(num_cols):
                        if i == first_col_index:
                            col_widths[i] = first_col_width
                        elif i in short_numeric_cols:
                            col_widths[i] = short_col_width
                        elif i in long_content_cols:
                            col_widths[i] = long_col_width
                        else:
                            col_widths[i] = regular_col_width

                    table = Table(data, colWidths=col_widths)

                    table_style = [
                        ('FONTNAME', (0, 0), (-1, 0), font_name_bold),
                        ('FONTNAME', (0, 1), (-1, -1), font_name),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5B5FC7')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTSIZE', (0, 0), (-1, 0), 11),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('FONTSIZE', (0, 1), (-1, -1), 10),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E7E7F7')]),
                    ]

                    # Apply left alignment and top valignment for long content columns and first column
                    for col_index in long_content_cols:
                        table_style.append(('ALIGN', (col_index, 1), (col_index, -1), 'LEFT'))
                        table_style.append(('VALIGN', (col_index, 1), (col_index, -1), 'TOP'))
                    
                    # First column (names) should also be left-aligned and top-aligned
                    table_style.append(('ALIGN', (first_col_index, 1), (first_col_index, -1), 'LEFT'))
                    table_style.append(('VALIGN', (first_col_index, 1), (first_col_index, -1), 'TOP'))

                    table.setStyle(TableStyle(table_style))
                    story.append(table)

                    footer_style = ParagraphStyle(
                        'Footer',
                        fontName=font_name,
                        parent=styles['Normal'],
                        fontSize=8,
                        spaceAfter=12,
                        alignment=1,
                        textColor=colors.grey
                    )

                    story.append(Spacer(1, 20))
                    footer_text = f"Nəticələr sayı: {len(pdf_df)} | Yaradılma tarixi: {pd.Timestamp.now().strftime('%d.%m.%Y %H:%M')}"
                    footer_para = Paragraph(footer_text, footer_style)
                    story.append(footer_para)

                    doc.build(story)
                    output.seek(0)
                    return output

                except Exception as e:
                    st.error(f"PDF yaradılarkən xəta: {e}")
                    return None


            # Replace the existing download button section with this updated version:

            # Create download buttons (only if duplicates are resolved)
            if allow_download:
                original_filename = uploaded_file.name
                trimmed_year = original_filename[:5]

                def trim_until_variant(s):
                    lower_s = s.lower()
                    idx = lower_s.find("variant")
                    if idx != -1:
                        return s[:idx]
                    else:
                        return s

                if selected_assignments:
                    first_filter = str(selected_assignments[0])
                    filter_part = trim_until_variant(first_filter)
                else:
                    filter_part = "unfiltered"
                

                excel_filename = f"{trimmed_year}{filter_part}.xlsx"
                pdf_filename = f"{trimmed_year}{filter_part}_dik.pdf"
                pdf_landscape_filename = f"{trimmed_year}{filter_part}.pdf"


                title_cell = title_cell[:-15] + filter_part

                excel_data = to_excel(final_filtered_df, title_cell)
                pdf_data = to_pdf(final_filtered_df, title_cell)
                pdf_landscape_data = to_pdf_landscape(final_filtered_df, title_cell)
                
                # Generate filenames

                
                # Create three columns for the buttons
                col1, col2, col3 = st.columns(3)
                
                # Excel button
                if excel_data:
                    with col1:
                        st.download_button(
                            label="📊 Excel faylını yüklə",
                            data=excel_data,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                # PDF Portrait button
                if pdf_data:
                    with col2:
                        st.download_button(
                            label="📄 PDF (Dikinə)",
                            data=pdf_data,
                            file_name=pdf_filename,
                            mime="application/pdf"
                        )
                
                # PDF Landscape button
                if pdf_landscape_data:
                    with col3:
                        st.download_button(
                            label="📄 PDF (Üfüqi)",
                            data=pdf_landscape_data,
                            file_name=pdf_landscape_filename,
                            mime="application/pdf"
                        )
                
                # Show warning if any format failed to generate
                if not excel_data:
                    st.warning("Excel faylı yaradıla bilmədi")
                if not pdf_data:
                    st.warning("PDF (Dik) faylı yaradıla bilmədi")
                if not pdf_landscape_data:
                    st.warning("PDF (Üfüqi) faylı yaradıla bilmədi")
            else:
                st.error("❌ Yükləmək olmaz. Təkrarları aradan qaldırın.")